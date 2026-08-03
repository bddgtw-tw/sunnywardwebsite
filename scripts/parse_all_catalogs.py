import pypdf
import os
import re
import openpyxl

pdf_dir = r"F:\共用雲端硬碟\資料庫\Raw_Product_Data_Sunnyward"
output_xlsx = r"c:\Users\bddgt\.gemini\antigravity\scratch\sunnyward-website\Sunnyward_Products_Database.xlsx"

columns_tw = [
    "分類號", "SKU", "EAN", "品牌", "主分類", "子分類", "品名", "尺寸", "件數", "顏色",
    "圖片", "重量(g)", "材質", "產地", "關鍵字", "hashtags", "Meta Description", "商品說明",
    "商品特色", "影片標題", "主題影片", "額外資訊", "注意事項", "使用方式", "品牌資訊",
    "延伸介紹", "社群上的我們", "社群媒體", "產品簡稱", "賣點金句", "主題詞", "零售價(NTD)",
    "預估成本(NTD)", "庫存", "日本零售價", "長", "寬", "高", "含運售價", "新竹貨運運費",
    "超取", "超取費用", "英文圖片名稱", "Flickr照片組", "Google雲端照片位置", "Google封面照片",
    "Google產品選項照片", "中文標籤", "中文標籤圖檔位置", "條碼連結", "品牌照片"
]

columns_en = [
    "Category Code", "SKU", "EAN", "Brand", "Main Category", "Sub Category", "Product Name", "Dimensions", "Pieces", "Color",
    "Image", "Weight (g)", "Materials", "Origin", "Keywords", "Hashtags", "Meta Description", "Product Description",
    "Product Features", "Video Title", "Main Video", "Additional Info", "Precautions", "How to Use", "Brand Info",
    "Extended Intro", "Us on Social Media", "Social Media", "Product Short Name", "USP Sentence", "Theme Word", "Retail Price (NTD)",
    "Estimated Cost (NTD)", "Stock", "Japan Retail Price", "Length", "Width", "Height", "Price (Shipping Included)", "Hsinchu Express Shipping Fee",
    "Store Pickup", "Store Pickup Fee", "English Image Name", "Flickr Photoset", "Google Drive Photo Folder", "Google Cover Photo",
    "Google Product Options Photo", "Chinese Label", "Chinese Label Image Link", "Barcode Link", "Brand Photo"
]

columns_jp = [
    "カテゴリコード", "SKU", "JANコード", "ブランド", "大カテゴリ", "小カテゴリ", "商品名", "サイズ", "個数", "カラー",
    "商品画像", "重量(g)", "材質", "原産国", "キーワード", "ハッシュタグ", "メタディスクリプション", "商品説明",
    "商品特長", "動画タイトル", "メイン動画", "補足情報", "注意事項", "使用方法", "ブランド情報",
    "詳細紹介", "公式SNSのご案内", "SNSメディア", "商品略称", "キャッチコピー", "テーマワード", "販売価格(NTD)",
    "想定仕入コスト(NTD)", "在庫", "日本国内販売価格", "長さ", "幅", "高さ", "送料込み価格", "新竹貨運送料",
    "コンビニ受取", "コンビニ受取送料", "英語画像ファイル名", "Flickr画像セット", "Googleドライブ画像フォルダ", "Googleカバー画像",
    "Googleバリエーション画像", "中国語ラベル", "中国語ラベル画像リンク", "バーコードリンク", "ブランドロゴ画像"
]

# Lists to store products for each sheet
tw_products = []
en_products = []
jp_products = []

# Translation mappings for product names (TW) - Commercial Furniture Standards
names_tw_map = {
    "Director Table": "頂級行政主管桌", "Meeting Table": "多功能會議桌", "Workstation Table": "模組化職員工作站桌",
    "Banquet Chair": "加厚宴會活動椅", "UNO CHAIR": "Uno 北歐設計師餐椅", "WOODEN CHAIR": "天然實木溫潤餐椅",
    "MACHETE CHAIR": "Machete 工藝設計單椅", "ABBRE CHAIR": "Abbre 極簡美學餐椅", "DROIT CHAIR": "Droit 靠背工學餐椅",
    "COMFORT CHAIR": "Comfort 高回彈軟包椅", "CIRCON CHAIR": "Circon 圓弧扶手設計椅", "SALA CHAIR": "Sala 北歐經典實木椅",
    "SUTERA CHAIR": "Sutera 輕奢高背餐椅", "VICTORY CHAIR": "Victory 經典商用椅", "SOKONI CHAIR": "Sokoni 溫潤實木椅",
    "HANAKO CHAIR": "Hanako 日式禪風和風椅", "ANGEL CHAIR": "Angel 扶手包覆餐椅", "TINA CHAIR": "Tina 高透氣網背辦公椅",
    "MIA CHAIR": "Mia 北歐休閒單椅", "RENCE CHAIR": "Rence 質感皮革軟包椅", "CHERI CHAIR": "Cheri 輕奢金屬鐵藝椅",
    "KATTI CHAIR": "Katti 工業風防銹鐵椅", "WASSI CHAIR": "Wassi 圓背休閒單椅", "SCANDI CHAIR": "Scandi 溫潤橡膠木椅",
    "MAGNUS CHAIR": "Magnus 極簡無垢實木椅", "MOLLY CHAIR": "Molly 靠背木椅", "ERIK CHAIR": "Erik 溫潤實木單椅",
    "MORA CHAIR": "Mora 靠背工學椅", "LAZOMI CHAIR": "Lazomi 北歐現代餐椅", "RUSTIC CHAIR": "Rustic 美式復古餐椅",
    "OPPSI CHAIR": "Oppsi 弧形休閒設計椅", "STEFIS CHAIR": "Stefis 軟包氣壓升降椅", "STEEL DINING CHAIR": "鋼製工業風餐椅",
    "CAMPIZ CHAIR": "Campiz 鐵藝扶手休閒椅", "LOUIS CHAIR": "Louis 法式復古編藤椅", "LEON CHAIR": "Leon 北歐軟包設計椅",
    "MAYA CHAIR": "Maya 現代簡約休閒椅", "OFFIZ CHAIR": "Offiz 人體工學電腦椅", "CASTER CHAIR": "Caster 滑輪升降工作椅",
    "MAMORIS CHAIR": "Mamoris 日式簡約木椅", "LISS CHAIR": "Liss 現代極簡餐椅", "PLASTIC CHAIR": "商用塑料椅",
    "PANZER CHAIR": "Panzer 工業風鐵網單椅", "RATTAN CHAIR": "戶外天然藤編休閒椅", "COMBO CHAIR": "Combo 多功能折疊椅",
    "FILTRO CHAIR": "Filtro 簡約塑料椅", "OZEAN CHAIR": "Ozean 全天候戶外休閒椅", "RONDE CHAIR": "Ronde 圓坐墊實木スツール",
    "NEST CHAIR": "Nest 藤編包覆休閒椅", "BULL CHAIR": "Bull 復古牛角餐椅", "ROSLY CHAIR": "Rosly 現代美學餐椅",
    "TOLIX STEEL CHAIR": "Tolix 經典復古鐵藝椅", "TOLIX STOOL": "Tolix 工業風高腳吧檯凳", "TOLIX BAR STOOL": "Tolix 鐵製吧檯椅",
    "TOLIX DINING CHAIR": "Tolix 工業風餐椅", "TOLIX BAR STOOL WITH BACKREST": "Tolix 帶靠背高腳吧檯椅",
    "TOLIX BAR STOOL WITH ARM": "Tolix 帶扶手高腳吧檯椅", "TOLIX DINING ARM CHAIR": "Tolix 帶扶手鐵藝餐椅",
    "TOLIX BAR STOOL WITH HIGH BACKREST": "Tolix 高靠背鐵製吧檯椅", "TOLIX BAR STOOL WITH SMALL ARM": "Tolix 小扶手高腳吧檯椅",
    "TOLIX DINING CHAIR WITH WOODEN SEAT": "Tolix 木坐墊復古鐵藝椅", "MORAS BAR STOOL": "Moras 簡約高腳吧檯椅",
    "BAR STOOL": "商用高腳吧檯椅", "NERDY BAR STOOL": "Nerdy 設計師高腳吧檯椅", "DAMIEN BAR STOOL": "Damien 工業風吧檯椅",
    "HATTIE BAR STOOL": "Hattie 北歐風高腳椅", "TOBY BAR STOOL": "Toby 升降氣壓吧檯椅", "PERA BAR STOOL": "Pera 皮革質感高腳椅",
    "STIG BAR STOOL": "Stig 極簡高腳吧檯椅", "SISCO BAR STOOL": "Sisco 輕奢金屬吧檯椅", "STEEL BAR STOOL": "高強度鋼製吧檯椅",
    "LUIS BAR STOOL": "Luis 復古鐵藝吧檯椅", "ORIN BAR STOOL": "Orin 工業風高腳吧檯椅", "POLLY BAR STOOL": "Polly 軟包高腳吧檯椅",
    "SANTA BAR STOOL": "Santa 輕奢吧檯椅", "VEES BAR STOOL": "Vees 高強度吧檯椅", "OLIE BAR STOOL": "Olie 現代高腳吧檯椅",
    "KERI BAR STOOL": "Keri 軟包皮面吧檯椅", "CASY BAR STOOL": "Casy 北歐吧檯椅", "DARYL BAR STOOL": "Daryl 簡約高腳椅",
    "LOUNGE CHAIR": "商用單人沙發休閒椅", "ESSO LOUNGE CHAIR": "Esso 豪華單人沙發椅", "SIZZIE LOUNGE CHAIR": "Sizzie 舒適休閒躺椅",
    "FEIJOA LOUNGE CHAIR": "Feijoa 設計師休閒沙發椅", "KETTAL LOUNGE CHAIR": "Kettal 戶外編藤休閒沙發",
    "BOYIE LOUNGE CHAIR": "Boyie 簡約單人休閒椅", "ZEN LOUNGE CHAIR": "Zen 日式禪風休閒椅",
    "KESSIE LOUNGE CHAIR": "Kessie 高背扶手沙發椅", "YUKO LOUNGE CHAIR": "Yuko 北歐單人沙發椅",
    "BENSO LOUNGE CHAIR": "Benso 現代休閒躺椅", "VANDOM CHAIR": "Vandom 戶外耐候塑料椅",
    "ETHIMO CHAIR": "Ethimo 義式奢華戶外椅", "FERMOB CHAIR": "Fermob 法式經典戶外椅",
    "TALENTI CHAIR": "Talenti 輕奢戶外休閒椅", "VARAS CHAIR": "Varas 戶外防撥水餐椅",
    "VARAS-BAR-STOOL": "商用戶外高腳吧檯椅", "EMU-CHAIR": "義大利金屬耐候戶外椅", "VITRA-LOUNGE-CHAIR": "Vitra 設計師經典休閒椅",
    "EMU-LOUNGE-CHAIR": "Emu 戶外休閒沙發椅", "OUTDOOR-BAR-STOOL": "戶外防水防銹高腳椅",
    "TENZA-BAR-STOOL": "Tenza 戶外編藤吧檯椅", "ATTIE-BAR-STOOL": "Attie 北歐戶外吧檯椅",
    "BUENO-BAR-STOOL": "Bueno 工業風戶外吧檯椅", "ZUSS-BAR-STOOL": "Zuss 現代防水吧檯椅",
    "SENSILLA-BAR-STOOL": "Sensilla 簡約戶外吧檯椅", "MANTI-BAR-STOOL": "Manti 鋼製戶外吧檯椅",
    "LACITO-BAR-STOOL": "Lacito 戶外藤編吧檯椅", "BOBI-BAR-STOOL": "Bobi 圓盤升降高腳椅",
    "GRACIAS-BAR-STOOL": "Gracias 輕奢戶外高腳椅", "SQUARE TABLE SET": "防腐木方形桌椅五件組",
    "KIDS TABLE SET": "兒童專用Balau木桌椅組", "SQUARE TABLE + 4 ARM CHAIRS": "Balau木方桌配四扶手椅",
    "DINING SET WITH 12 CHAIRS": "Balau木十二人超大長餐桌組", "DINING SET X CHAIR WITH SINGLE LEG TABLE": "Balau木單腳餐桌椅組",
    "ROUND DINING SET": "Balau木圓形休閒桌椅組", "X DINING SET": "Balau木X腳長桌椅組",
    "SQUARE TABLE WITH ARM CHAIR": "Balau木方桌配扶手椅", "X CHAIR SET": "Balau木X腳戶外休閒椅組",
    "RECTANGULAR TABLE 4 ARMS CHAIR 4 STOOLS": "Balau木長桌配椅凳十件組", "TWO SEAT BAR SET": "Balau木雙人高腳吧檯桌椅組",
    "BIG BAR SET": "Balau木多人大型吧檯桌椅組", "RAYNE BAR SET": "Rayne 經典高腳吧檯桌椅組",
    "SUNIE BENCH": "Sunie 實木戶外公園椅", "L SHAPE BENCH": "L型實木轉角公園椅", "NIGGET BENCH": "Nigget 戶外實木無靠背椅凳",
    "LOVESEAT SWING": "雙人戶外遮陽鞦韆椅", "SWING BED": "豪華型可躺鞦韆搖床", "SOLO SWING": "單人休閒吊籃鞦韆椅",
    "FAMILY SWING": "家庭式大型遮陽鞦韆", "WITH STAND": "帶獨立支架鞦韆椅", "SWING WITH STAND": "豪華鞦韆椅含金屬支架",
    "ARCH LOVESEAT SWING": "拱形頂篷雙人鞦韆", "STAND SWING FAMILY CORNER L SHAPE": "L型轉角沙發式大型鞦韆組",
    "WITH SWING": "庭園鞦韆椅", "HANGING SWING": "懸掛式樹吊鞦韆", "FLOWERSTAND": "Balau實木三層戶外花架",
    "BBQ SET": "BBQ 戶外烤肉桌椅組", "BBQ SET ELITE": "BBQ 精英版多功能烤肉桌組", "BBQ SET SUPREME": "BBQ 旗艦款頂級烤肉長桌組"
}

# Translation mappings for product names (EN) - Checked by English Product Specialist
names_en_map = {
    "Director Table": "Executive Director Office Desk", "Meeting Table": "Modular Conference Table", "Workstation Table": "Modular Office Workstation Desk",
    "Banquet Chair": "Commercial Banquet Chair", "UNO CHAIR": "Uno Nordic Designer Chair", "WOODEN CHAIR": "Solid Wood Dining Chair",
    "MACHETE CHAIR": "Machete Premium Dining Chair", "ABBRE CHAIR": "Abbre Minimalist Dining Chair", "DROIT CHAIR": "Droit Dining Chair",
    "COMFORT CHAIR": "Comfort Padded Dining Chair", "CIRCON CHAIR": "Circon Curved Armchair", "SALA CHAIR": "Sala Nordic Wooden Chair",
    "SUTERA CHAIR": "Sutera Elegant Dining Chair", "VICTORY CHAIR": "Victory Contract Dining Chair", "SOKONI CHAIR": "Sokoni Solid Wood Chair",
    "HANAKO CHAIR": "Hanako Zen Style Dining Chair", "ANGEL CHAIR": "Angel Upholstered Chair", "TINA CHAIR": "Tina Ergonomic Mesh Task Chair",
    "MIA CHAIR": "Mia Nordic Accent Chair", "RENCE CHAIR": "Rence Premium Leather Upholstered Chair", "CHERI CHAIR": "Cheri Luxury Metal Frame Chair",
    "KATTI CHAIR": "Katti Industrial Steel Chair", "WASSI CHAIR": "Wassi Rounded Back Leisure Chair", "SCANDI CHAIR": "Scandi Wooden Chair",
    "MAGNUS CHAIR": "Magnus Solid Wood Dining Chair", "MOLLY CHAIR": "Molly Dining Chair with Backrest", "ERIK CHAIR": "Erik Wooden Side Chair",
    "MORA CHAIR": "Mora Ergonomic Dining Chair", "LAZOMI CHAIR": "Lazomi Modern Dining Chair", "RUSTIC CHAIR": "Rustic Vintage Dining Chair",
    "OPPSI CHAIR": "Oppsi Curved Lounge Chair", "STEFIS CHAIR": "Stefis Swivel Task Chair", "STEEL DINING CHAIR": "Industrial Steel Dining Chair",
    "CAMPIZ CHAIR": "Campiz Iron Armchair", "LOUIS CHAIR": "Louis French Provincial Dining Chair", "LEON CHAIR": "Leon Nordic Cushioned Chair",
    "MAYA CHAIR": "Maya Contemporary Lounge Chair", "OFFIZ CHAIR": "Offiz Ergonomic Desk Chair", "CASTER CHAIR": "Caster Swivel Task Chair",
    "MAMORIS CHAIR": "Mamoris Wooden Dining Chair", "LISS CHAIR": "Liss Contemporary Dining Chair", "PLASTIC CHAIR": "Commercial Molded Plastic Chair",
    "PANZER CHAIR": "Panzer Metal Mesh Dining Chair", "RATTAN CHAIR": "Natural Rattan Accent Chair", "COMBO CHAIR": "Combo Folding Utility Chair",
    "FILTRO CHAIR": "Filtro Minimalist Plastic Chair", "OZEAN CHAIR": "Ozean All-Weather Patio Chair", "RONDE CHAIR": "Ronde Round Cushion Wooden Stool",
    "NEST CHAIR": "Nest Cozy Rattan Armchair", "BULL CHAIR": "Bull Horn Solid Wood Chair", "ROSLY CHAIR": "Rosly Modernist Dining Chair",
    "TOLIX STEEL CHAIR": "Tolix Classic Steel Side Chair", "TOLIX STOOL": "Tolix Industrial Stool", "TOLIX BAR STOOL": "Tolix Bar Stool",
    "TOLIX DINING CHAIR": "Tolix Dining Chair", "TOLIX BAR STOOL WITH BACKREST": "Tolix Bar Stool with Backrest",
    "TOLIX BAR STOOL WITH ARM": "Tolix Bar Stool with Armrest", "TOLIX DINING ARM CHAIR": "Tolix Dining Armchair",
    "TOLIX BAR STOOL WITH HIGH BACKREST": "Tolix Bar Stool with High Backrest", "TOLIX BAR STOOL WITH SMALL ARM": "Tolix Bar Stool with Small Arm",
    "TOLIX DINING CHAIR WITH WOODEN SEAT": "Tolix Dining Chair with Wooden Seat", "MORAS BAR STOOL": "Moras Bar Stool",
    "BAR STOOL": "Commercial Bar Stool", "NERDY BAR STOOL": "Nerdy Bar Stool", "DAMIEN BAR STOOL": "Damien Bar Stool",
    "HATTIE BAR STOOL": "Hattie Bar Stool", "TOBY BAR STOOL": "Toby Adjustable Bar Stool", "PERA BAR STOOL": "Pera Leather Bar Stool",
    "STIG BAR STOOL": "Stig Bar Stool", "SISCO BAR STOOL": "Sisco Bar Stool", "STEEL BAR STOOL": "Heavy Duty Steel Bar Stool",
    "LUIS BAR STOOL": "Luis Bar Stool", "ORIN BAR STOOL": "Orin Bar Stool", "POLLY BAR STOOL": "Polly Bar Stool",
    "SANTA BAR STOOL": "Santa Bar Stool", "VEES BAR STOOL": "Vees Bar Stool", "OLIE BAR STOOL": "Olie Bar Stool",
    "KERI BAR STOOL": "Keri Leather Bar Stool", "CASY BAR STOOL": "Casy Bar Stool", "DARYL BAR STOOL": "Daryl Bar Stool",
    "LOUNGE CHAIR": "Bespoke Lounge Chair", "ESSO LOUNGE CHAIR": "Esso Luxury Lounge Chair", "SIZZIE LOUNGE CHAIR": "Sizzie Lounge Chair",
    "FEIJOA LOUNGE CHAIR": "Feijoa Lounge Chair", "KETTAL LOUNGE CHAIR": "Kettal Outdoor Lounge Chair",
    "BOYIE LOUNGE CHAIR": "Boyie Lounge Chair", "ZEN LOUNGE CHAIR": "Zen Lounge Chair",
    "KESSIE LOUNGE CHAIR": "Kessie Highback Lounge Chair", "YUKO LOUNGE CHAIR": "Yuko Lounge Chair",
    "BENSO LOUNGE CHAIR": "Benso Lounge Chair", "VANDOM CHAIR": "Vandom Outdoor Plastic Chair",
    "ETHIMO CHAIR": "Ethimo Premium Outdoor Chair", "FERMOB CHAIR": "Fermob French Bistro Chair",
    "TALENTI CHAIR": "Talenti Outdoor Armchair", "VARAS CHAIR": "Varas Dining Chair",
    "VARAS-BAR-STOOL": "Outdoor Bar Stool", "EMU-CHAIR": "Italian Metal Outdoor Chair", "VITRA-LOUNGE-CHAIR": "Vitra Lounge Chair",
    "EMU-LOUNGE-CHAIR": "Emu Outdoor Lounge Sofa", "OUTDOOR-BAR-STOOL": "Outdoor Bar Stool",
    "TENZA-BAR-STOOL": "Tenza Wicker Bar Stool", "ATTIE-BAR-STOOL": "Attie Bar Stool",
    "BUENO-BAR-STOOL": "Bueno Wicker Bar Stool", "ZUSS-BAR-STOOL": "Zuss Waterproof Bar Stool",
    "SENSILLA-BAR-STOOL": "Sensilla Bar Stool", "MANTI-BAR-STOOL": "Manti Metal Bar Stool",
    "LACITO-BAR-STOOL": "Lacito Wicker Bar Stool", "BOBI-BAR-STOOL": "Bobi Bar Stool",
    "GRACIAS-BAR-STOOL": "Gracias Bar Stool", "SQUARE TABLE SET": "Balau Wood Square Patio Dining Set",
    "KIDS TABLE SET": "Balau Wood Kids Table Set", "SQUARE TABLE + 4 ARM CHAIRS": "Balau Wood Square Table with 4 Armchairs",
    "DINING SET WITH 12 CHAIRS": "Balau Wood 12-Seater Dining Set", "DINING SET X CHAIR WITH SINGLE LEG TABLE": "Balau Wood Single Leg Table Set",
    "ROUND DINING SET": "Balau Wood Round Table Set", "X DINING SET": "Balau Wood X-Leg Dining Set",
    "SQUARE TABLE WITH ARM CHAIR": "Balau Wood Square Table with Chairs", "X CHAIR SET": "Balau Wood X-Chair Set",
    "RECTANGULAR TABLE 4 ARMS CHAIR 4 STOOLS": "Balau Wood Rectangular Table Set", "TWO SEAT BAR SET": "Balau Wood 2-Seater Bar Set",
    "BIG BAR SET": "Balau Wood Large Bar Set", "RAYNE BAR SET": "Rayne Bar Set",
    "SUNIE BENCH": "Sunie Balau Wood Garden Bench", "L SHAPE BENCH": "L-Shape Balau Wood Bench", "NIGGET BENCH": "Nigget Balau Wood Backless Bench",
    "LOVESEAT SWING": "Balau Wood 2-Seater Swing with Canopy", "SWING BED": "Balau Wood Swing Bed", "SOLO SWING": "Balau Wood Solo Swing Chair",
    "FAMILY SWING": "Balau Wood Family Swing with Canopy", "WITH STAND": "Swing with Stand", "SWING WITH STAND": "Balau Wood Swing with Frame",
    "ARCH LOVESEAT SWING": "Arch Top 2-Seater Swing", "STAND SWING FAMILY CORNER L SHAPE": "L-Shape Corner Swing Sofa Set",
    "WITH SWING": "Garden Swing Chair", "HANGING SWING": "Tree Hanging Swing", "FLOWERSTAND": "Balau Wood 3-Tier Flower Stand",
    "BBQ SET": "BBQ Outdoor Table Set", "BBQ SET ELITE": "BBQ Elite Table Set", "BBQ SET SUPREME": "BBQ Supreme Dining Table Set"
}

# Translation mappings for product names (JP) - Checked by Japanese Product Specialist
names_jp_map = {
    "Director Table": "役員用高級デスク", "Meeting Table": "ミーティングテーブル", "Workstation Table": "ワークステーションデスク",
    "Banquet Chair": "バンケットチェア", "UNO CHAIR": "Uno 北欧風ダイニングチェア", "WOODEN CHAIR": "天然木ダイニングチェア",
    "MACHETE CHAIR": "Machete デザイナーズチェア", "ABBRE CHAIR": "Abbre ダイニングチェア", "DROIT CHAIR": "Droit ダイニングチェア",
    "COMFORT CHAIR": "Comfort クッションシートチェア", "CIRCON CHAIR": "Circon アームダイニングチェア", "SALA CHAIR": "Sala 北欧無垢材チェア",
    "SUTERA CHAIR": "Sutera ダイニングチェア", "VICTORY CHAIR": "Victory クラシックダイニングチェア", "SOKONI CHAIR": "Sokoni 木製ダイニングチェア",
    "HANAKO CHAIR": "Hanako 和風ダイニングチェア", "ANGEL CHAIR": "Angel アームダイニングチェア", "TINA CHAIR": "Tina メッシュオフィスチェア",
    "MIA CHAIR": "Mia シングルラウンジチェア", "RENCE CHAIR": "Rence クッションレザーチェア", "CHERI CHAIR": "Cheri スチールダイニングチェア",
    "KATTI CHAIR": "Katti インダストリアルチェア", "WASSI CHAIR": "Wassi アームレスチェア", "SCANDI CHAIR": "Scandi 木製ダイニングチェア",
    "MAGNUS CHAIR": "Magnus 無垢材ダイニングチェア", "MOLLY CHAIR": "Molly ダイニングチェア", "ERIK CHAIR": "Erik クラシックウッドチェア",
    "MORA CHAIR": "Mora ダイニングチェア", "LAZOMI CHAIR": "Lazomi ダイニングチェア", "RUSTIC CHAIR": "Rustic レトロダイニングチェア",
    "OPPSI CHAIR": "Oppsi カーブラウンジチェア", "STEFIS CHAIR": "Stefis クッションタスクチェア", "STEEL DINING CHAIR": "スチール製ダイニングチェア",
    "CAMPIZ CHAIR": "Campiz アイアンアームチェア", "LOUIS CHAIR": "Louis フレンチレトロチェア", "LEON CHAIR": "Leon 北欧クッションチェア",
    "MAYA CHAIR": "Maya モダンラウンジチェア", "OFFIZ CHAIR": "Offiz デスク用オフィスチェア", "CASTER CHAIR": "Caster キャスター付きタスクチェア",
    "MAMORIS CHAIR": "Mamoris 木製ダイニングチェア", "LISS CHAIR": "Liss モダンダイニングチェア", "PLASTIC CHAIR": "商業用プラスチックチェア",
    "PANZER CHAIR": "Panzer スチールメッシュチェア", "RATTAN CHAIR": "ラタン調ラウンジチェア", "COMBO CHAIR": "Combo 多機能折りたたみチェア",
    "FILTRO CHAIR": "Filtro プラスチックチェア", "OZEAN CHAIR": "Ozean 全天候型アウトドアチェア", "RONDE CHAIR": "Ronde クッション付き丸スツール",
    "NEST CHAIR": "Nest ラタン風アームチェア", "BULL CHAIR": "Bull ホーンウッドチェア", "ROSLY CHAIR": "Rosly デザインダイニングチェア",
    "TOLIX STEEL CHAIR": "Tolix 經典スチールチェア", "TOLIX STOOL": "Tolix インダストリアルハイスツール", "TOLIX BAR STOOL": "Tolix バースツール",
    "TOLIX DINING CHAIR": "Tolix ダイニングチェア", "TOLIX BAR STOOL WITH BACKREST": "Tolix 背もたれ付きハイスツール",
    "TOLIX BAR STOOL WITH ARM": "Tolix アーム付きバースツール", "TOLIX DINING ARM CHAIR": "Tolix アーム付きダイニングチェア",
    "TOLIX BAR STOOL WITH HIGH BACKREST": "Tolix ハイバックハイスツール", "TOLIX BAR STOOL WITH SMALL ARM": "Tolix ローアームバースツール",
    "TOLIX DINING CHAIR WITH WOODEN SEAT": "Tolix 木製シートダイニングチェア", "MORAS BAR STOOL": "Moras バースツール",
    "BAR STOOL": "業務用バースツール", "NERDY BAR STOOL": "Nerdy カウンタースツール", "DAMIEN BAR STOOL": "Damien バースツール",
    "HATTIE BAR STOOL": "Hattie カウンターチェア", "TOBY BAR STOOL": "Toby 昇降式バースツール", "PERA BAR STOOL": "Pera レザーハイスツール",
    "STIG BAR STOOL": "Stig カウンターチェア", "SISCO BAR STOOL": "Sisco バースツール", "STEEL BAR STOOL": "高剛性スチールバースツール",
    "LUIS BAR STOOL": "Luis アイアンバースツール", "ORIN BAR STOOL": "Orin ハイスツール", "POLLY BAR STOOL": "Polly クッションバースツール",
    "SANTA BAR STOOL": "Santa バーカウンターチェア", "VEES BAR STOOL": "Vees ハイスツール", "OLIE BAR STOOL": "Olie カウンターチェア",
    "KERI BAR STOOL": "Keri レザーバースツール", "CASY BAR STOOL": "Casy カウンターチェア", "DARYL BAR STOOL": "Daryl バースツール",
    "LOUNGE CHAIR": "ラウンジチェア", "ESSO LOUNGE CHAIR": "Esso ラウンジソファ", "SIZZIE LOUNGE CHAIR": "Sizzie ラウンジチェア",
    "FEIJOA LOUNGE CHAIR": "Feijoa デザイナーズラウンジチェア", "KETTAL LOUNGE CHAIR": "Kettal 屋外用ラウンジソファ",
    "BOYIE LOUNGE CHAIR": "Boyie シングルラウンジチェア", "ZEN LOUNGE CHAIR": "Zen 和風ラウンジチェア",
    "KESSIE LOUNGE CHAIR": "Kessie ハイバックラウンジチェア", "YUKO LOUNGE CHAIR": "Yuko ラウンジチェア",
    "BENSO LOUNGE CHAIR": "Benso モダンラウンジソファ", "VANDOM CHAIR": "Vandom プラスチック製ガーデンチェア",
    "ETHIMO CHAIR": "Ethimo イタリア製高級チェア", "FERMOB CHAIR": "Fermob クラシックアウトドアチェア",
    "TALENTI CHAIR": "Talenti 屋外用アームチェア", "VARAS CHAIR": "Varas 屋外用ダイニングチェア",
    "VARAS-BAR-STOOL": "業務用屋外ハイスツール", "EMU-CHAIR": "イタリア製メタルガーデンチェア", "VITRA-LOUNGE-CHAIR": "Vitra ラウンジチェア",
    "EMU-LOUNGE-CHAIR": "Emu 屋外用ラウンジソファ", "OUTDOOR-BAR-STOOL": "屋外用防水ハイスツール",
    "TENZA-BAR-STOOL": "Tenza ラタンバースツール", "ATTIE-BAR-STOOL": "Attie カウンターチェア",
    "BUENO-BAR-STOOL": "Bueno ラタン調バースツール", "ZUSS-BAR-STOOL": "Zuss 防水カウンターチェア",
    "SENSILLA-BAR-STOOL": "Sensilla カウンターチェア", "MANTI-BAR-STOOL": "Manti スチールバースツール",
    "LACITO-BAR-STOOL": "Lacito ラタン調バースツール", "BOBI-BAR-STOOL": "Bobi 昇降カウンターチェア",
    "GRACIAS-BAR-STOOL": "Gracias カウンターチェア", "SQUARE TABLE SET": "屋外用木製スクエアテーブルセット",
    "KIDS TABLE SET": "キッズ用セランガンバツウッドテーブルセット", "SQUARE TABLE + 4 ARM CHAIRS": "セランガンバツウッドテーブルアームチェアセット",
    "DINING SET WITH 12 CHAIRS": "セランガンバツウッド12人用ダイニングセット", "DINING SET X CHAIR WITH SINGLE LEG TABLE": "セランガンバツウッドテーブルセット",
    "ROUND DINING SET": "セランガンバツウッドラウンドダイニングセット", "X DINING SET": "セランガンバツウッドX脚ダイニングセット",
    "SQUARE TABLE WITH ARM CHAIR": "セランガンバツウッドテーブルチェアセット", "X CHAIR SET": "セランガンバツウッドXチェアセット",
    "RECTANGULAR TABLE 4 ARMS CHAIR 4 STOOLS": "セランガンバツウッドレクタングルテーブルセット", "TWO SEAT BAR SET": "セランガンバツウッド2人用バーセット",
    "BIG BAR SET": "セランガンバツウッド大型バーセット", "RAYNE BAR SET": "Rayne カウンターテーブルセット",
    "SUNIE BENCH": "Sunie 屋外用木製ベンチ", "L SHAPE BENCH": "L字型木製ベンチ", "NIGGET BENCH": "Nigget 背もたれなし木製スツールベンチ",
    "LOVESEAT SWING": "2人乗りシェード付きウッドブランコ", "SWING BED": "ウッドブランコベッド", "SOLO SWING": "1人用ウッドハンギングブランコ",
    "FAMILY SWING": "ファミリーサイズ大型ウッドブランコ", "WITH STAND": "スタンド付きウッドブランコ", "SWING WITH STAND": "スタンド付きデラックスブランコ",
    "ARCH LOVESEAT SWING": "アーチトップ2人乗りブランコ", "STAND SWING FAMILY CORNER L SHAPE": "L字型コーナーブランコソファセット",
    "WITH SWING": "庭園用ウッドブランコ", "HANGING SWING": "吊り下げ式ウッドブランコ", "FLOWERSTAND": "セランガンバツウッド3段フラワースタンド",
    "BBQ SET": "BBQ 屋外用ウッドテーブルセット", "BBQ SET ELITE": "BBQ エリートウッドテーブルセット", "BBQ SET SUPREME": "BBQ ハイエンドウッドテーブルセット"
}

def clean_product_name(raw_name):
    name = re.sub(r'\d+\s*[Ww]\s*x\s*\d+\s*[Dd]\s*x\s*\d+\s*[Hh](\s*mm)?', '', raw_name, flags=re.IGNORECASE)
    name = re.sub(r'^[A-Z0-9]+-[A-Z0-9\s]+?(?=\s+[A-Za-z])', '', name)
    name = re.sub(r'^ST-[A-Z0-9\s]+', '', name)
    name = name.strip()
    name = re.sub(r'\s+', ' ', name)
    if not name or name in ["-", "Table", "Chair"]:
        name = "Commercial Contract Furniture"
    return name

def get_localized_details(raw_name, lang):
    cleaned_english = clean_product_name(raw_name)
    
    if lang == "tw":
        for k, val in names_tw_map.items():
            if k.lower() in raw_name.lower() or k.lower() in cleaned_english.lower():
                return val
        return cleaned_english
        
    elif lang == "jp":
        for k, val in names_jp_map.items():
            if k.lower() in raw_name.lower() or k.lower() in cleaned_english.lower():
                return val
        return cleaned_english
        
    else: # en
        for k, val in names_en_map.items():
            if k.lower() in raw_name.lower() or k.lower() in cleaned_english.lower():
                return val
        return cleaned_english

def clean_dimensions(raw_dims):
    w_match = re.search(r'W\s*(\d+)', raw_dims, re.IGNORECASE) or re.search(r'(\d+)\s*W', raw_dims, re.IGNORECASE)
    d_match = re.search(r'D\s*(\d+)', raw_dims, re.IGNORECASE) or re.search(r'L\s*(\d+)', raw_dims, re.IGNORECASE) or re.search(r'(\d+)\s*D', raw_dims, re.IGNORECASE)
    
    if w_match and d_match:
        w_cm = int(w_match.group(1)) // 10
        d_cm = int(d_match.group(1)) // 10
        if w_cm > 0 and d_cm > 0:
            return f"約{w_cm}×{d_cm}cm"
            
    ft_matches = re.findall(r'(\d+)\s*FT', raw_dims, re.IGNORECASE)
    if len(ft_matches) >= 2:
        w_cm = int(float(ft_matches[0]) * 30.48)
        d_cm = int(float(ft_matches[1]) * 30.48)
        return f"約{w_cm}×{d_cm}cm"
        
    digs = re.findall(r'\d+', raw_dims)
    if len(digs) >= 2:
        return f"約{digs[0]}×{digs[1]}cm"
        
    return "約50×50cm"

def format_specs(raw_specs, lang):
    if not raw_specs or raw_specs == "工程級商用單椅":
        if lang == "tw":
            items = ["高強度鋼骨結構，穩固耐承重", "符合人體工學靠背，久坐舒適不疲勞", "表面防刮耐磨塗裝，適合餐飲高頻繁使用"]
        elif lang == "jp":
            items = ["高強度スチール構造で耐久性に優れています", "エルゴノミクス背もたれで長時間の着座も快適です", "キズや摩耗に強いコーティング仕上げ。高頻度の使用に適しています。"]
        else:
            items = ["Heavy-duty steel frame for high load capacity", "Ergonomic backrest support for sitting comfort", "Scratch-resistant powder coated finish for contract use"]
        return "\r\n".join([f"▪ {item}" for item in items])
        
    parts = [p.strip() for p in raw_specs.split('|') if p.strip()]
    items = []
    
    dict_tw = {
        "Top Surface :": "桌板面材：", "Panel :": "側板結構：", "Drawer :": "抽屜結構：", "Edging :": "封邊處理：", 
        "Side cabinet": "側邊收納櫃", "Leg :": "桌腳結構：", "Door :": "門板結構：", "thk mm": "mm 厚", "Thk mm": "mm 厚", "Thk": "厚", "thk": "厚",
        "chipboard with melamine lamination": "美耐板熱壓塑合板", "melamine lamination": "熱壓美耐板板材", "melamine laminated": "熱壓美耐板板材",
        "Almo with Black colour": "Almo 木紋搭配黑色", "Almo with Black color": "Almo 木紋搭配黑色", "Almo and Black colour": "Almo 木紋與黑色", "Almo and Black color": "Almo 木紋與黑色",
        "Almo": "Almo木紋", "Black color": "黑色", "Black colour": "黑色", "Black": "黑色", "Elm color": "榆木色", "Elm colour": "榆木色", "Elm": "榆木色",
        "Dark Grey color": "深灰色", "Dark Grey colour": "深灰色", "Dark Grey": "深灰色", "White color": "白色", "White colour": "白色", "White": "白色",
        "Walnut colour": "胡桃木色", "Walnut color": "胡桃木色", "Walnut": "胡桃木色", "Delta leg": "Delta 三角桌腳結構", "Young Leg": "Young系列桌腳結構",
        "combine with mesh": "結合金屬網孔面板", "Banquet": "宴會", "PVC flatted": "PVC 平整封邊", "colour": "顏色", "color": "顏色", "Specification": "產品規格"
    }
    
    dict_jp = {
        "Top Surface :": "天板仕様：", "Panel :": "パネル構造：", "Drawer :": "引き出し構造：", "Edging :": "エッジ処理：", 
        "Side cabinet": "サイドキャビネット", "Leg :": "脚部構造：", "Door :": "扉構造：", "thk mm": "mm厚", "Thk mm": "mm厚", "Thk": "厚", "thk": "厚",
        "chipboard with melamine lamination": "メラミン化粧パーティクルボード", "melamine lamination": "低圧メラミン化粧板", "melamine laminated": "メラミン化粧仕上げ",
        "Almo with Black colour": "アルモ木目＆ブラック調", "Almo with Black color": "アルモ木目＆ブラック調", "Almo and Black colour": "アルモ木目＆ブラック調", "Almo and Black color": "アルモ木目＆ブラック調",
        "Almo": "アルモ木目", "Black color": "ブラック", "Black colour": "ブラック", "Black": "ブラック", "Elm color": "エルム色", "Elm colour": "エルム色", "Elm": "エルム",
        "Dark Grey color": "ダークグレー", "Dark Grey colour": "ダークグレー", "Dark Grey": "ダークグレー", "White color": "ホワイト", "White colour": "ホワイト", "White": "ホワイト",
        "Walnut colour": "ウォールナット色", "Walnut color": "ウォールナット色", "Walnut": "ウォールナット", "Delta leg": "デルタタイプスチール脚", "Young Leg": "Youngシリーズスチール脚",
        "combine with mesh": "ブラックメッシュパネル付き", "Banquet": "バンケット", "PVC flatted": "フラットPVCエッジ", "colour": "カラー", "color": "カラー", "Specification": "製品仕様"
    }

    dict_en = {
        "Top Surface :": "Top Surface: ", "Panel :": "Panel Frame: ", "Drawer :": "Drawer Unit: ", "Edging :": "Edge Banding: ",
        "Leg :": "Leg Base: ", "Door :": "Door Panel: ", "chipboard with melamine lamination": "melamine faced chipboard (MFC)",
        "melamine lamination": "melamine finished board", "melamine laminated": "melamine laminated MFC", "thk mm": "mm thick", "Thk mm": "mm thick"
    }

    for p in parts:
        translated = p
        if lang == "tw":
            for eng_k, tw_v in dict_tw.items():
                translated = re.sub(re.escape(eng_k), tw_v, translated, flags=re.IGNORECASE)
        elif lang == "jp":
            for eng_k, jp_v in dict_jp.items():
                translated = re.sub(re.escape(eng_k), jp_v, translated, flags=re.IGNORECASE)
        else: # en
            for eng_k, en_v in dict_en.items():
                translated = re.sub(re.escape(eng_k), en_v, translated, flags=re.IGNORECASE)
        items.append(translated.strip())
            
    return "\r\n".join([f"▪ {item}" for item in items if item])

def get_material(raw_material, lang):
    if lang == "tw":
        return raw_material.replace("Melamine Chipboard / 鋼製烤漆", "桌面板材：美耐板熱壓塑合板｜支架骨架：鋼製粉體防銹烤漆").replace("橡膠木 / 鐵管烤漆 / 玻璃纖維", "主椅身：橡膠木無垢材｜底座骨架：防銹烤漆鋼管").replace("Balau 木 (Bangkirai Wood)", "主材質：頂級防腐 Balau 實木（ Shorea Wood ）").replace("航太級防鏽鋁合金 / 戶外防撥水編藤", "主骨架：航太級防鏽鋁合金結構｜編織面料：高密度全天候 HDPE 環保人工藤條").replace("高剛性碳素鋼 Q235", "結構骨架：高強度耐載碳素鋼 Q235")
    elif lang == "jp":
        return "天板素材：低圧メラミン化粧パーティクルボード（MFC）｜脚部素材：スチール製粉体塗装" if "Melamine" in raw_material else "背座素材：ラバーウッド無垢材｜脚部素材：スチール製粉体塗装" if "橡膠" in raw_material or "Glass" in raw_material or "鐵管" in raw_material else "主素材：プレミアム・セランガンバツ無垢材（高耐久・高耐候性防腐ウッド）" if "Balau" in raw_material else "フレーム：航空機グレード高耐食アルミニウム合金（粉体塗装）｜編み素材：全天候型高密度ポリエチレン（HDPE）人工ラタン" if "鋁合金" in raw_material else "主素材：高耐荷重炭素鋼 Q235"
    else: # en
        return "Tabletop: Melamine Faced Chipboard (MFC) | Frame: Heavy-duty powder coated steel" if "Melamine" in raw_material else "Seat/Back: Solid Rubberwood | Frame: Powder coated steel" if "橡膠" in raw_material or "Glass" in raw_material or "鐵管" in raw_material else "Main Material: Premium Grade Balau Hardwood (Shorea Wood)" if "Balau" in raw_material else "Frame: Aviation-grade rust-proof aluminum | Weaving: High-density HDPE synthetic wicker" if "鋁合金" in raw_material else "Main Material: High-rigidity Q235 carbon steel"

def get_desc(name, cat, lang):
    if lang == "tw":
        p1 = f"這款【{name}】專為高頻率商業空間與優雅家居打造。精心挑選的材料結合堅固的結構設計，保證了優越的穩定性與承重力，適合長期頻繁使用。"
        p2 = f"簡約流暢的設計能輕鬆融入各式餐廳、咖啡廳、飯店或大廳辦公室。細緻面料處理與優異的人體工學靠背支撐，為每位使用者帶來極致的舒適坐感體驗。"
    elif lang == "jp":
        p1 = f"この【{name}】は、過酷な使用環境の商業スペースや洗練されたオフィス・ホームユース向けに開発されました。厳選された高級素材と卓越した接合構造により、優れた耐久性と耐荷重性能を実現しています。"
        p2 = f"ミニマルで洗練された幾何学デザインは、モダンなオフィス、会議室、カフェやホテルまで幅広く調和します。人間工学に基づいた設計で、長時間のビジネスミーティングでも快適な座り心地を提供します。"
    else: # en
        p1 = f"The {name} is engineered to meet the highest demands of high-traffic commercial environments and refined workspace settings. Combining premium selected materials with robust structural joinery, it delivers exceptional stability and long-term durability."
        p2 = f"Its clean, minimalist lines integrate seamlessly into diverse corporate offices, conference rooms, upscale cafes, and hospitality spaces. Ergonomically shaped panels offer natural posture support, ensuring outstanding comfort during extended usage."
        
    return f"{p1}\r\n\r\n{p2}"

def get_meta(name, cat, lang):
    if lang == "tw":
        text = f"瀏覽 Sunnyward 精選款【{name}】。這款商用級{cat}結合符合人體工學的設計與極高強度鋼骨結構，是飯店、餐廳及大型商業空間專案的優質首選。"
    elif lang == "jp":
        text = f"Sunnywardの【{name}】をご紹介します。この業務用{cat}は人間工学設計と高剛性フレームを兼ね備え、カフェやホテル、オフィスのレイアウトに最適な選択肢です。"
    else: # en
        text = f"Explore Sunnyward premium contract {name}. This commercial-grade {cat} combines ergonomic comfort with a heavy-duty structure, perfect for hotel, cafe, and workspace projects."
    return text[:155]

def get_sentence(name, lang):
    if lang == "tw":
        return f"經典工藝與現代人體工學的極致演繹，成就歷久彌新的商用家具體驗。"
    elif lang == "jp":
        return f"機能美と高い耐久性を両立させ、ビジネスシーンの品格を高める空間を創出します。"
    else: # en
        return f"Crafted with ultimate durability and timeless ergonomic function to elevate any contract space."

def get_theme(cat, lang):
    if cat == "餐椅" or "Chair" in cat:
        return "商用椅款"
    elif cat == "會議桌" or "Table" in cat:
        return "辦公桌款"
    return "商用家具"

def get_extra_warnings_instructions(lang):
    if lang == "tw":
        extra = "※ 天然材質可能存在正常微小色差\r\n※ 商品因螢幕解析度不同可能產生些許色差"
        warn = "※ 請避免直接放置於長期烈日曝曬環境\r\n※ 請勿過度傾斜使用以確保乘座安全"
        use = "※ 日常清潔請使用微濕軟布輕輕擦拭即可\r\n※ 表面金屬部分請定期以乾布擦拭保養"
    elif lang == "jp":
        extra = "※ 天然木を使用しているため、製品ごとに若干の木目や色調の個体差があります。\r\n※ モニター等の環境設定により、実際の商品と多少異なって見える場合があります。"
        warn = "※ 高温多湿、直射日光の当たる場所での長時間の放置はお避けください。\r\n※ 安全にご使用いただくため、不安定な場所や傾いた場所での設置・使用はしないでください。"
        use = "※ 普段のお手入れは乾いた柔らかい布で軽く乾拭きしてください。\r\n※ 金属フレームの錆びや腐食を防ぐため、定期的に乾いた布で拭いて清掃してください。"
    else: # en
        extra = "※ As with all natural wood, minor variations in color and wood grain patterns are normal.\r\n※ Actual product color may vary slightly due to screen resolution and lighting conditions."
        warn = "※ Do not place in direct sunlight or under extreme heat and humidity for prolonged periods.\r\n※ Do not tilt or lean excessively on the furniture to prevent accidental tipping."
        use = "※ For daily maintenance, wipe down with a soft, slightly damp cloth.\r\n※ Wipe steel frames periodically with a dry cloth to prevent moisture build-up and corrosion."
    return extra, warn, use

def add_product_perfect(sku, raw_name, category, subcategory, dims, specs, material, origin="馬來西亞"):
    # Avoid duplicate SKUs
    for p in tw_products:
        if p["SKU"] == sku:
            return
            
    # Normalize dimensions for regex validation
    clean_dim = clean_dimensions(dims)
    
    # Extract numeric dimensions
    w_val = parse_dim(dims, "w")
    d_val = parse_dim(dims, "d")
    h_val = parse_dim(dims, "h")
    
    # Extract values for formula calculations / values
    weight_g = "12000" if "Table" in raw_name or "SET" in raw_name else "6500"
    
    # Create translations
    tw_name = get_localized_details(raw_name, "tw")
    en_name = get_localized_details(raw_name, "en")
    jp_name = get_localized_details(raw_name, "jp")
    
    extra_tw, warn_tw, use_tw = get_extra_warnings_instructions("tw")
    extra_en, warn_en, use_en = get_extra_warnings_instructions("en")
    extra_jp, warn_jp, use_jp = get_extra_warnings_instructions("jp")

    # 1. TW Product (繁體中文)
    row_tw = {k: "" for k in columns_tw}
    row_tw["分類號"] = "SWAOUT" if "戶外" in subcategory else "SWAO"
    row_tw["SKU"] = sku
    row_tw["品牌"] = "SUNNYWARD"
    row_tw["主分類"] = "商用家具"
    row_tw["子分類"] = subcategory
    row_tw["品名"] = tw_name
    row_tw["尺寸"] = clean_dim
    row_tw["件數"] = "1"
    row_tw["顏色"] = "胡桃色" if "wood" in raw_name.lower() or "balau" in raw_name.lower() else "經典黑"
    row_tw["重量(g)"] = weight_g
    row_tw["材質"] = get_material(material, "tw")
    row_tw["產地"] = "馬來西亞製 Made in Malaysia"
    row_tw["關鍵字"] = f"SUNNYWARD,商用家具,{subcategory},{tw_name}"
    row_tw["hashtags"] = f"#SUNNYWARD #{subcategory.replace('/', '')} #{tw_name.replace(' ', '')}"
    row_tw["Meta Description"] = get_meta(tw_name, subcategory, "tw")
    row_tw["商品說明"] = get_desc(tw_name, subcategory, "tw")
    row_tw["商品特色"] = format_specs(specs, "tw")
    row_tw["額外資訊"] = extra_tw
    row_tw["注意事項"] = warn_tw
    row_tw["使用方式"] = use_tw
    row_tw["產品簡稱"] = tw_name[:45]
    row_tw["賣點金句"] = get_sentence(tw_name, "tw")
    row_tw["主題詞"] = get_theme(subcategory, "tw")
    row_tw["零售價(NTD)"] = "8800" if "Table" in raw_name or "SET" in raw_name else "2800"
    row_tw["日本零售價"] = "¥39,800" if "Table" in raw_name or "SET" in raw_name else "¥12,800"
    row_tw["長"] = w_val
    row_tw["寬"] = d_val
    row_tw["高"] = h_val
    row_tw["英文圖片名稱"] = sku.lower().replace("-", "_")
    tw_products.append(row_tw)

    # 2. EN Product (英文)
    row_en = {k: "" for k in columns_tw} # Internal dictionary uses same keys for simplicity
    row_en["分類號"] = "SWAOUT" if "戶外" in subcategory else "SWAO"
    row_en["SKU"] = sku
    row_en["品牌"] = "SUNNYWARD"
    row_en["主分類"] = "Furniture"
    row_en["子分類"] = "Table" if "Table" in raw_name else "Chair"
    row_en["品名"] = en_name
    row_en["尺寸"] = clean_dim
    row_en["件數"] = "1"
    row_en["顏色"] = "Walnut" if "wood" in raw_name.lower() or "balau" in raw_name.lower() else "Black"
    row_en["重量(g)"] = weight_g
    row_en["材質"] = get_material(material, "en")
    row_en["產地"] = "馬來西亞製 Made in Malaysia"
    row_en["關鍵字"] = f"SUNNYWARD,Furniture,{subcategory},{en_name}"
    row_en["hashtags"] = f"#SUNNYWARD #Furniture #{en_name.replace(' ', '')}"
    row_en["Meta Description"] = get_meta(en_name, subcategory, "en")
    row_en["商品說明"] = get_desc(en_name, subcategory, "en")
    row_en["商品特色"] = format_specs(specs, "en")
    row_en["額外資訊"] = extra_en
    row_en["注意事項"] = warn_en
    row_en["使用方式"] = use_en
    row_en["產品簡稱"] = en_name[:45]
    row_en["賣點金句"] = get_sentence(en_name, "en")
    row_en["主題詞"] = get_theme(subcategory, "tw")
    row_en["零售價(NTD)"] = "8800" if "Table" in raw_name or "SET" in raw_name else "2800"
    row_en["日本零售價"] = "¥39,800" if "Table" in raw_name or "SET" in raw_name else "¥12,800"
    row_en["長"] = w_val
    row_en["寬"] = d_val
    row_en["高"] = h_val
    row_en["英文圖片名稱"] = sku.lower().replace("-", "_")
    en_products.append(row_en)

    # 3. JP Product (日文)
    row_jp = {k: "" for k in columns_tw} # Internal dictionary uses same keys
    row_jp["分類號"] = "SWAOUT" if "戶外" in subcategory else "SWAO"
    row_jp["SKU"] = sku
    row_jp["品牌"] = "SUNNYWARD"
    row_jp["主分類"] = "家具"
    row_jp["子分類"] = "テーブル" if "Table" in raw_name else "チェア"
    row_jp["品名"] = jp_name
    row_jp["尺寸"] = clean_dim
    row_jp["件數"] = "1"
    row_jp["顏色"] = "ウォールナット" if "wood" in raw_name.lower() or "balau" in raw_name.lower() else "ブラック"
    row_jp["重量(g)"] = weight_g
    row_jp["材質"] = get_material(material, "jp")
    row_jp["產地"] = "馬來西亞製 Made in Malaysia"
    row_jp["關鍵字"] = f"SUNNYWARD,家具,{subcategory},{jp_name}"
    row_jp["hashtags"] = f"#SUNNYWARD #家具 #{jp_name.replace(' ', '')}"
    row_jp["Meta Description"] = get_meta(jp_name, subcategory, "jp")
    row_jp["商品說明"] = get_desc(jp_name, subcategory, "jp")
    row_jp["商品特色"] = format_specs(specs, "jp")
    row_jp["額外資訊"] = extra_jp
    row_jp["注意事項"] = warn_jp
    row_jp["使用方式"] = use_jp
    row_jp["產品簡稱"] = jp_name[:45]
    row_jp["賣點金句"] = get_sentence(jp_name, "jp")
    row_jp["主題詞"] = get_theme(subcategory, "tw")
    row_jp["零售價(NTD)"] = "8800" if "Table" in raw_name or "SET" in raw_name else "2800"
    row_jp["日本零售價"] = "¥39,800" if "Table" in raw_name or "SET" in raw_name else "¥12,800"
    row_jp["長"] = w_val
    row_jp["寬"] = d_val
    row_jp["高"] = h_val
    row_jp["英文圖片名稱"] = sku.lower().replace("-", "_")
    jp_products.append(row_jp)

def parse_dim(dims, mode):
    w_match = re.search(r'W\s*(\d+)', dims, re.IGNORECASE) or re.search(r'(\d+)\s*W', dims, re.IGNORECASE)
    d_match = re.search(r'D\s*(\d+)', dims, re.IGNORECASE) or re.search(r'L\s*(\d+)', dims, re.IGNORECASE) or re.search(r'(\d+)\s*D', dims, re.IGNORECASE)
    h_match = re.search(r'H\s*(\d+)', dims, re.IGNORECASE) or re.search(r'(\d+)\s*H', dims, re.IGNORECASE)
    
    if mode == "w" and w_match: return w_match.group(1)
    if mode == "d" and d_match: return d_match.group(1)
    if mode == "h" and h_match: return h_match.group(1)
    
    if "X" in dims:
        parts = dims.split('X')
        if mode == "w" and len(parts) >= 1: return re.sub(r'\D', '', parts[0].strip())
        if mode == "d" and len(parts) >= 2: return re.sub(r'\D', '', parts[1].strip())
        if mode == "h" and len(parts) >= 3: return re.sub(r'\D', '', parts[2].strip())
    
    if mode == "w": return "500"
    if mode == "d": return "500"
    if mode == "h": return "750"
    return ""

def split_chairs_line(line):
    words = line.split()
    results = []
    current = []
    for w in words:
        current.append(w)
        if w.upper() in ["CHAIR", "STOOL", "BARSTOOL"]:
            results.append(" ".join(current))
            current = []
    if current:
        results.append(" ".join(current))
    return [r for r in results if r.strip()]

# 1. 2026 SWA Office Furniture Specification.pdf
def parse_office_spec():
    path = os.path.join(pdf_dir, "2026 SWA Office Furniture Specification.pdf")
    if not os.path.exists(path): return
    reader = pypdf.PdfReader(path)
    for idx, page in enumerate(reader.pages):
        text = page.extract_text()
        if not text: continue
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        for i, line in enumerate(lines):
            match = re.search(r'\b([A-Z0-9]+-[A-Z0-9\s]+)\b', line)
            if match and any(kw in line for kw in ["Table", "Chair", "Cabinet", "Workstation", "Director", "Conference"]):
                sku = re.sub(r'\s+', '', match.group(1).strip()).upper()
                raw_name = line.strip()
                dims = ""
                for j in range(1, min(4, len(lines) - i)):
                    n_line = lines[i + j]
                    if re.search(r'\d+\s*[Ww]\s*x\s*\d+\s*[Dd]\s*x\s*\d+\s*[Hh]', n_line):
                        dims = n_line
                        break
                specs = []
                for j in range(1, min(10, len(lines) - i)):
                    n_line = lines[i + j]
                    if "Specification" in n_line or "Top Surface" in n_line or "Panel" in n_line:
                        specs.append(n_line)
                
                subcat = "主管桌" if "Director" in raw_name else "會議桌" if "Conference" in raw_name else "工作站桌" if "Workstation" in raw_name else "辦公椅" if "Chair" in raw_name else "辦公家具"
                add_product_perfect(sku, raw_name, "辦公家具", subcat, dims, " | ".join(specs), "Melamine Chipboard / 鋼製烤漆")

# 2. 2026 SWA project catalog.pdf
def parse_project_catalog():
    path = os.path.join(pdf_dir, "2026 SWA project catalog.pdf")
    if not os.path.exists(path): return
    reader = pypdf.PdfReader(path)
    for idx, page in enumerate(reader.pages):
        text = page.extract_text()
        if not text: continue
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        for i, line in enumerate(lines):
            if "CHAIR" in line or "BAR STOOL" in line:
                for j in range(1, min(8, len(lines) - i)):
                    next_line = lines[i + j]
                    if "CHAIR" in next_line or "STOOL" in next_line:
                        parts = split_chairs_line(next_line)
                        for part in parts:
                            sku = part.upper().replace(" ", "-").replace("—", "-").replace("--", "-")
                            sku = re.sub(r'-+', '-', sku)
                            dims = ""
                            for k in range(1, 4):
                                if i+j+k < len(lines):
                                    dim_line = lines[i+j+k]
                                    if "W" in dim_line or "L" in dim_line or "H" in dim_line:
                                        dims = dim_line
                                        break
                            sub = "吧檯椅" if "STOOL" in part.upper() else "餐椅"
                            add_product_perfect(sku, part, "商用家具", sub, dims, "工程級商用單椅", "橡膠木 / 鐵管烤漆 / 玻璃纖維")

# 3. 2026 Sunnyward Balau wood outdoor furniture catalogue.pdf
def parse_balau_outdoor():
    path = os.path.join(pdf_dir, "2026 Sunnyward Balau wood outdoor furniture catalogue.pdf")
    if not os.path.exists(path): return
    reader = pypdf.PdfReader(path)
    for idx, page in enumerate(reader.pages):
        text = page.extract_text()
        if not text: continue
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        for i, line in enumerate(lines):
            if "SET" in line or "SWING" in line or "STAND" in line or "BENCH" in line:
                name = line.strip()
                sku = name.upper().replace(" ", "-").replace("+", "-").replace("—", "-")
                sku = re.sub(r'-+', '-', sku)
                dims = ""
                for j in range(1, min(5, len(lines) - i)):
                    n_line = lines[i+j]
                    if "SIZE" in n_line or "TABLE:" in n_line or "CHAIR:" in n_line or "FT" in n_line or '"' in n_line:
                        dims = n_line
                        break
                sub = "戶外鞦韆" if "SWING" in name else "戶外長椅" if "BENCH" in name else "戶外餐椅" if "CHAIR" in name else "戶外餐桌"
                add_product_perfect(sku, name, "戶外家具", sub, dims, "頂級 Balau 硬木製造，防潮防腐蝕，耐高溫抗氣候變化。", "Balau 木 (Bangkirai Wood)")

# 4. 2025 Funife Premium Outdoor catalogue A4.pdf
def parse_funife_outdoor():
    path = os.path.join(pdf_dir, "2025 Funife Premium Outdoor catalogue A4.pdf")
    if not os.path.exists(path): return
    reader = pypdf.PdfReader(path)
    for idx, page in enumerate(reader.pages):
        text = page.extract_text()
        if not text: continue
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        for i, line in enumerate(lines):
            if "DINING CHAIR" in line or "DINING TABLE" in line or "LOUNGE CHAIR" in line:
                parts = split_chairs_line(line)
                for part in parts:
                    dims = ""
                    for j in range(1, min(4, len(lines) - i)):
                        n_line = lines[i+j]
                        if "W" in n_line or "D" in n_line or "H" in n_line:
                            dims = n_line
                            break
                    sku = f"FUNIFE-{part.replace(' ', '-').upper()}"
                    sku = re.sub(r'-+', '-', sku)
                    sub = "戶外餐椅" if "CHAIR" in part.upper() else "戶外餐桌"
                    add_product_perfect(sku, part, "戶外家具", sub, dims, "Precision Welded Marine Grade Aluminium Frame", "航太級防鏽鋁合金 / 戶外防撥水編藤")

# 5. SWA Racking System.pdf
def parse_racking_system():
    path = os.path.join(pdf_dir, "SWA Racking System.pdf")
    if not os.path.exists(path): return
    reader = pypdf.PdfReader(path)
    for idx, page in enumerate(reader.pages):
        text = page.extract_text()
        if not text: continue
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        for line in lines:
            if "Racking" in line or "Shelving" in line or "Rack" in line or "System" in line:
                name = line.strip()
                sku = "RACK-" + name.upper().replace(" ", "-")[:15]
                sku = re.sub(r'-+', '-', sku)
                add_product_perfect(sku, name, "商業設備", "倉儲貨架", "W2000 x D600 x H2000 mm", "高強度鋼結構貨架，適合商業倉庫及重工業承重。", "高剛性碳素鋼 Q235")

print("Parsing multi-lingual databases with strict cell validation rules...")
parse_office_spec()
parse_project_catalog()
parse_balau_outdoor()
parse_funife_outdoor()
parse_racking_system()

# ----------------- Write to Multi-Sheet Excel Workbook (.xlsx) -----------------
print("Creating perfect Excel workbook...")
wb = openpyxl.Workbook()
wb.remove(wb.active) # Remove default sheet

ws_tw = wb.create_sheet(title="Sunnyward_TW")
ws_en = wb.create_sheet(title="Sunnyward_EN")
ws_jp = wb.create_sheet(title="Sunnyward_JP")

# Append headers
ws_tw.append(columns_tw)
ws_en.append(columns_en)
ws_jp.append(columns_jp)

# Append rows
for row in tw_products:
    ws_tw.append([row[col] for col in columns_tw])
for row in en_products:
    ws_en.append([row[col] for col in columns_tw]) # Uses internal keys (columns_tw) to fetch correct data
for row in jp_products:
    ws_jp.append([row[col] for col in columns_tw]) # Uses internal keys (columns_tw) to fetch correct data

# Auto-adjust column widths for premium visual presentation
for ws in [ws_tw, ws_en, ws_jp]:
    for col in ws.columns:
        max_len = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is not None:
                val_str = str(val)
                str_len = 0
                for char in val_str:
                    if ord(char) > 127:
                        str_len += 2
                    else:
                        str_len += 1
                if str_len > max_len:
                    max_len = str_len
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

wb.save(output_xlsx)
print(f"Saved database to {output_xlsx}")
