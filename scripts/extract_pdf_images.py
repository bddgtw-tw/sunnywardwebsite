import os
import re
import fitz  # PyMuPDF
import openpyxl

pdf_dir = r"F:\共用雲端硬碟\資料庫\Raw_Product_Data_Sunnyward"
base_image_dir = r"c:\Users\bddgt\.gemini\antigravity\scratch\sunnyward-website\Product_Images"
output_xlsx = r"c:\Users\bddgt\.gemini\antigravity\scratch\sunnyward-website\Sunnyward_Products_Database.xlsx"

# 1. Load the database to get the map of SKU to category folders
sku_to_folder = {}
wb = openpyxl.load_workbook(output_xlsx)
ws = wb["Sunnyward_TW"]
headers = [cell.value for cell in ws[1]]
sku_idx = headers.index("SKU") + 1
subcat_idx = headers.index("子分類") + 1

for row in range(2, ws.max_row + 1):
    sku = ws.cell(row=row, column=sku_idx).value
    subcat = ws.cell(row=row, column=subcat_idx).value
    if not sku:
        continue
    sku = sku.strip().upper()
    
    if subcat in ["主管桌", "會議桌", "工作站桌", "辦公椅", "辦公家具"]:
        cat_folder = "01_Office_Furniture"
    elif subcat in ["餐椅", "吧檯椅", "商用沙發"]:
        cat_folder = "02_Commercial_Furniture"
    elif subcat in ["戶外鞦韆", "戶外長椅", "戶外餐椅", "戶外餐桌"]:
        cat_folder = "03_Outdoor_Furniture"
    else:
        cat_folder = "04_Commercial_Equipment"
        
    sku_to_folder[sku] = os.path.join(base_image_dir, cat_folder, sku)

# List of PDFs to process
pdf_files = [f for f in os.listdir(pdf_dir) if f.endswith(".pdf")]

print("Starting image extraction from PDFs and organizing by SKU...")

for pdf_file in pdf_files:
    pdf_path = os.path.join(pdf_dir, pdf_file)
    print(f"\nProcessing PDF: {pdf_file}")
    
    doc = fitz.open(pdf_path)
    
    for page_num in range(len(doc)):
        page = doc[page_num]
        text = page.get_text()
        
        # Find which SKUs from our database appear on this page
        page_skus = []
        for sku in sku_to_folder.keys():
            # Check for exact word matches of the SKU or substrings
            # Simple check: is SKU string inside text?
            if sku in text.upper():
                page_skus.append(sku)
            else:
                # Some SKUs might be written with spaces in the PDF, e.g. "ST-LN 01" in text vs "ST-LN01" in database
                spaced_sku = sku.replace("-", " ")
                if spaced_sku in text.upper():
                    page_skus.append(sku)
                elif sku.replace("-", "") in text.upper():
                    page_skus.append(sku)
        
        # If no SKU found in text, check if we can extract it using general patterns
        # e.g., ST-MN01, ST-LN01, etc.
        if not page_skus:
            # Look for general patterns in office furniture spec sheet page
            matches = re.findall(r'\b(ST-[A-Z0-9\-]+)\b', text.upper())
            for match in matches:
                clean_match = match.replace(" ", "")
                if clean_match in sku_to_folder:
                    page_skus.append(clean_match)
        
        # If still no SKUs are found, we check the catalog filename to guess
        # (e.g. if the page has images, but no text, or text is non-searchable images)
        # For SWA project catalog or Balau catalog, we can use the text on the page.
        
        # Extract images from this page
        image_list = page.get_images(full=True)
        if not image_list:
            continue
            
        if page_skus:
            print(f"  Page {page_num + 1}: Found SKUs {page_skus}. Extracting {len(image_list)} images...")
            for img_idx, img_info in enumerate(image_list):
                xref = img_info[0]
                base_image = doc.extract_image(xref)
                image_bytes = base_image["image"]
                image_ext = base_image["ext"]
                
                # Save to each SKU folder
                for sku in page_skus:
                    target_folder = sku_to_folder[sku]
                    os.makedirs(target_folder, exist_ok=True)
                    
                    # We can name the file like: {sku.lower()}_{page_num+1}_{img_idx+1}.{image_ext}
                    filename = f"{sku.lower()}_page{page_num+1}_img{img_idx+1}.{image_ext}"
                    dest_path = os.path.join(target_folder, filename)
                    
                    with open(dest_path, "wb") as img_file:
                        img_file.write(image_bytes)
        else:
            # If no SKUs are found but there are images on this page, let's log it or save to a temp folder
            # to make sure we don't miss anything.
            pass

print("\nImage extraction and organization complete!")
