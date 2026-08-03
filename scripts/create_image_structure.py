import os
import openpyxl

base_dir = r"c:\Users\bddgt\.gemini\antigravity\scratch\sunnyward-website\Product_Images"
output_xlsx = r"c:\Users\bddgt\.gemini\antigravity\scratch\sunnyward-website\Sunnyward_Products_Database.xlsx"

print("Reading database to create SKU-based subfolders...")
try:
    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb["Sunnyward_TW"]
    
    # Identify index of needed columns
    headers = [cell.value for cell in ws[1]]
    sku_idx = headers.index("SKU") + 1
    subcat_idx = headers.index("子分類") + 1
    
    count = 0
    for row in range(2, ws.max_row + 1):
        sku = ws.cell(row=row, column=sku_idx).value
        subcat = ws.cell(row=row, column=subcat_idx).value
        
        if not sku:
            continue
            
        # Clean SKU for safe folder name
        sku_folder_name = sku.strip().upper()
        
        # Decide category folder
        if subcat in ["主管桌", "會議桌", "工作站桌", "辦公椅", "辦公家具"]:
            category_folder = "01_Office_Furniture"
        elif subcat in ["餐椅", "吧檯椅", "商用沙發"]:
            category_folder = "02_Commercial_Furniture"
        elif subcat in ["戶外鞦韆", "戶外長椅", "戶外餐椅", "戶外餐桌"]:
            category_folder = "03_Outdoor_Furniture"
        else:
            category_folder = "04_Commercial_Equipment"
            
        # Create full folder path: base_dir / category_folder / SKU
        sku_path = os.path.join(base_dir, category_folder, sku_folder_name)
        os.makedirs(sku_path, exist_ok=True)
        count += 1
        
    print(f"Successfully created {count} SKU-based folders inside category folders.")
except Exception as e:
    print(f"Error during folder creation: {str(e)}")
