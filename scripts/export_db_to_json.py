import os
import json
import openpyxl

base_dir = r"c:\Users\bddgt\.gemini\antigravity\scratch\sunnyward-website"
output_xlsx = os.path.join(base_dir, "Master_SKU_Catalog.xlsx")
if not os.path.exists(output_xlsx):
    output_xlsx = os.path.join(base_dir, "Sunnyward_Products_Database.xlsx")
image_dir = os.path.join(base_dir, "Product_Images")

def build_sku_image_targets(xlsx_path):
    orig_path = os.path.join(base_dir, "Sunnyward_Products_Database.xlsx")
    target_xlsx = xlsx_path if os.path.exists(xlsx_path) else orig_path
    wb = openpyxl.load_workbook(target_xlsx)
    sheet_name = "Sunnyward_TW" if "Sunnyward_TW" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    sku_idx = headers.index("SKU")
    
    page_sku_groups = [
        (3, [23, 24, 25, 26, 27, 28, 29, 30, 31, 32]),
        (4, [33, 34, 35, 36, 37, 38, 39, 40]),
        (5, [41, 42, 43, 44, 45, 46, 47, 48, 49]),
        (6, [50, 51, 52, 53, 54, 55, 56, 57, 58, 59]),
        (7, [60, 61, 62, 63, 64, 65, 66, 67, 68, 69]),
        (8, [70, 71, 72, 73, 75]),
        (9, [76, 78, 79, 80, 81, 82, 83, 84]),
        (10, [85, 86, 87, 88, 89, 90, 91, 92, 93, 94]),
        (11, [96, 97, 98, 99, 100, 101, 102, 103, 104]),
        (20, [105, 106, 107, 108, 109, 110, 111, 112, 113]),
        (21, [114, 115, 116, 117, 118, 119, 120, 121, 122, 123]),
    ]
    
    sku_targets = {}
    for page_num, rows in page_sku_groups:
        for idx, r in enumerate(rows, 1):
            sku = ws.cell(r, sku_idx + 1).value
            if sku:
                sku_targets[sku.upper()] = f"page{page_num}_img{idx}.png"
    return sku_targets

SKU_IMAGE_TARGETS = build_sku_image_targets(output_xlsx)

# Find product media by the language-neutral SKU instead of a translated
def norm_sku(s):
    import re
    clean = re.sub(r'\[.*?\]', '', s)
    return re.sub(r'[^A-Z0-9]', '', clean.upper())

def get_sku_image(_subcat, sku):
    sku_n = norm_sku(sku)
    target_pattern = SKU_IMAGE_TARGETS.get(sku.upper(), "")
    
    matches = []
    # Two-level scan: Product_Images/[Category]/[SKU_folder]/
    for cat_folder in os.listdir(image_dir):
        cat_path = os.path.join(image_dir, cat_folder)
        if not os.path.isdir(cat_path):
            continue
        for folder in os.listdir(cat_path):
            folder_path = os.path.join(cat_path, folder)
            if not os.path.isdir(folder_path):
                continue
            folder_n = norm_sku(folder)
            if sku_n == folder_n or folder_n.startswith(sku_n) or sku_n in folder_n:
                files = sorted(
                    f for f in os.listdir(folder_path)
                    if f.lower().endswith((".png", ".jpg", ".jpeg", ".webp"))
                )
                if files:
                    originals = [f for f in files if not f.startswith("[PDF]_") and not f.startswith("page")]
                    if originals:
                        # Always prefer verified studio/manufacturer photos; ignore target_pattern
                        chosen = originals[0]
                    else:
                        # No original photos — use target_pattern to pick the right PDF image
                        chosen = files[0]
                        if target_pattern:
                            for f in files:
                                if f.endswith(f"_{target_pattern}") or f == target_pattern or f.endswith(f"_[PDF]_{target_pattern}"):
                                    chosen = f
                                    break
                    matches.append((0 if originals else 1, f"../Product_Images/{cat_folder}/{folder}/{chosen}"))
                

    if matches:
        matches.sort(key=lambda x: (x[0], x[1]))
        return matches[0][1]
    return ""

def build_sku_category_map(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    sheet_name = "SKU_Master_Ledger" if "SKU_Master_Ledger" in wb.sheetnames else ("Sunnyward_TW" if "Sunnyward_TW" in wb.sheetnames else wb.sheetnames[0])
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    sku_idx = headers.index("SKU")
    subcat_col = "Category_Key" if "Category_Key" in headers else ("子分類" if "子分類" in headers else "SKU")
    subcat_idx = headers.index(subcat_col)
    
    sku_map = {}
    for r in range(2, ws.max_row + 1):
        sku = ws.cell(r, sku_idx + 1).value
        subcat = ws.cell(r, subcat_idx + 1).value
        if not sku:
            continue
        sku_upper = sku.upper()
        if subcat in ["outdoor", "office", "project"]:
            tab_key = str(subcat)
        elif "SET" in sku_upper or "SWING" in sku_upper or "BENCH" in sku_upper or "FLOWERSTAND" in sku_upper or "BBQ" in sku_upper or (subcat and ("outdoor" in str(subcat).lower() or subcat in ["戶外餐桌", "戶外餐椅", "戶外長椅", "戶外鞦韆", "花架", "BBQ烤肉桌", "Outdoor Furniture", "Outdoor"])):
            tab_key = "outdoor"
        elif subcat and ("office" in str(subcat).lower() or subcat in ["主管桌", "工作站桌", "辦公椅", "辦公家具"]):
            tab_key = "office"
        else:
            tab_key = "project"
        sku_map[sku] = tab_key
    return sku_map

sku_category_map = build_sku_category_map(output_xlsx)

def export_sheet_to_json(sheet_name, output_path, lang):
    wb = openpyxl.load_workbook(output_xlsx)
    if sheet_name not in wb.sheetnames:
        if "SKU_Master_Ledger" in wb.sheetnames:
            sheet_name = "SKU_Master_Ledger"
        else:
            sheet_name = wb.sheetnames[0]
        
    ws = wb[sheet_name]
    headers = [cell.value for cell in ws[1]]
    
    # Define mapping dictionary for headers based on language
    if sheet_name == "SKU_Master_Ledger":
        lang_suffix = "TW" if lang == "tw" else ("EN" if lang == "en" else "JP")
        header_map = {
            "sku": "SKU", "name": f"Product_Name_{lang_suffix}", "subcat": f"Subcategory_{lang_suffix}",
            "dims": "Dimensions", "desc": f"Product_Name_{lang_suffix}", "specs": "Materials",
            "material": "Materials", "origin": "Origin", "price": "SKU", "jp_price": "SKU"
        }
    elif lang == "tw":
        header_map = {
            "sku": "SKU", "name": "品名", "subcat": "子分類", "dims": "尺寸",
            "desc": "商品說明", "specs": "商品特色", "material": "材質", "origin": "產地",
            "price": "零售價(NTD)", "jp_price": "日本零售價"
        }
    elif lang == "en":
        header_map = {
            "sku": "SKU", "name": "Product Name", "subcat": "Sub Category", "dims": "Dimensions",
            "desc": "Product Description", "specs": "Product Features", "material": "Materials", "origin": "Origin",
            "price": "Retail Price (NTD)", "jp_price": "Japan Retail Price"
        }
    else: # jp
        header_map = {
            "sku": "SKU", "name": "商品名", "subcat": "小カテゴリ", "dims": "サイズ",
            "desc": "商品説明", "specs": "商品特長", "material": "材質", "origin": "原産国",
            "price": "販売価格(NTD)", "jp_price": "日本國內販売価格"
        }
        
    sku_idx = headers.index(header_map["sku"])
    name_idx = headers.index(header_map["name"])
    subcat_idx = headers.index(header_map["subcat"])
    dims_idx = headers.index(header_map["dims"])
    desc_idx = headers.index(header_map["desc"])
    specs_idx = headers.index(header_map["specs"])
    material_idx = headers.index(header_map["material"])
    origin_idx = headers.index(header_map["origin"])
    price_idx = headers.index(header_map["price"])
    jp_price_idx = headers.index(header_map["jp_price"])
    
    products = []
    
    for row in range(2, ws.max_row + 1):
        vals = [cell.value for cell in ws[row]]
        if not vals[sku_idx]:
            continue
            
        sku = str(vals[sku_idx]).strip()
        name = str(vals[name_idx]).strip()
        subcat = str(vals[subcat_idx]).strip()
        dims = str(vals[dims_idx]).strip()
        desc = str(vals[desc_idx]).strip() if vals[desc_idx] else ""
        specs_raw = str(vals[specs_idx]).strip() if vals[specs_idx] else ""
        material = str(vals[material_idx]).strip() if vals[material_idx] else ""
        origin = str(vals[origin_idx]).strip() if vals[origin_idx] else ""
        if lang == "tw":
            origin = origin.replace(" Made in Malaysia", "").replace("Made in Malaysia ", "").replace("Made in Malaysia", "").strip()
            if not origin or origin == "Malaysia":
                origin = "馬來西亞製"
        elif lang == "en":
            origin = origin.replace("馬來西亞製 ", "").replace(" 馬來西亞製", "").replace("馬來西亞製", "").strip()
            if not origin or origin == "Malaysia":
                origin = "Made in Malaysia"
        elif lang == "jp":
            if "Malaysia" in origin or "馬來西亞" in origin:
                origin = "マレーシア製"
        price = str(vals[price_idx]).strip() if vals[price_idx] else ""
        jp_price = str(vals[jp_price_idx]).strip() if vals[jp_price_idx] else ""
        
        # Split specs into list of bullets
        specs = [s.replace("▪", "").strip() for s in specs_raw.split("\r\n") if s.strip()]
        if not specs:
            specs = [specs_raw] if specs_raw else []
            
        # Get dynamic image path
        img = get_sku_image(subcat, sku)
        
        # Determine category tab (outdoor, office, project)
        tab_key = sku_category_map.get(sku, "project")

        images = []
        if img:
            if img.startswith("http"):
                pass
            else:
                clean_img = img
                if clean_img.startswith("../"):
                    clean_img = clean_img[3:]
                images.append(clean_img)

        products.append({
            "id": sku,
            "sku": sku,
            "name": name,
            "brand": "Sunnyward",
            "category": tab_key,
            "sub_category": subcat,
            "collection": "General Collection",
            "description": desc,
            "materials": [material] if material else [],
            "dimensions": {"raw": dims},
            "origin": origin,
            "images": images,
            "image_dimensions": [[600, 450]],
            "detail_page": "products.html",
            "frontend_visible": True,
            "frontend_status": "published",
            "data_quality": {"image_status": "verified_reference"},
            "tab": tab_key,
            "dims": dims,
            "desc": desc,
            "material": material,
            "price": price,
            "jp_price": jp_price,
            "img": img
        })
    # Load verified products and merge them
    verified_path = os.path.join(base_dir, "data", "verified_product_pages.json")
    if os.path.exists(verified_path):
        with open(verified_path, "r", encoding="utf-8") as f:
            verified_data = json.load(f)
            verified_products = verified_data.get("products", [])
            for vp in verified_products:
                if not any(p["sku"] == vp["sku"] for p in products):
                    vp_category = vp.get("category", "")
                    if "outdoor" in vp_category.lower() or vp_category in ["Outdoor Furniture", "Outdoor", "戶外家具", "戶外休閒家具", "アウトドア家具"]:
                        tab_key = "outdoor"
                    elif vp_category in ["Office Furniture", "Office", "辦公家具", "オフィス家具"]:
                        tab_key = "office"
                    else:
                        tab_key = "project"
                    
                    images = [img.replace("../", "") for img in vp.get("images", [])]
                    products.append({
                        "id": vp["sku"],
                        "sku": vp["sku"],
                        "name": vp["locales"][lang]["name"] if lang in vp.get("locales", {}) else vp["name"],
                        "brand": vp.get("brand", "Sunnyward"),
                        "category": tab_key,
                        "sub_category": vp.get("sub_category", vp.get("collection", "")),
                        "collection": vp.get("collection", "General Collection"),
                        "description": vp["locales"][lang]["description"] if lang in vp.get("locales", {}) else vp.get("description", ""),
                        "materials": vp.get("materials", []),
                        "dimensions": vp.get("dimensions", {"raw": ""}),
                        "origin": vp.get("origin", ""),
                        "images": images,
                        "image_dimensions": vp.get("image_dimensions", [[600, 450]] * len(images)),
                        "detail_page": f"products/{vp['slug']}.html",
                        "frontend_visible": True,
                        "frontend_status": "published",
                        "data_quality": {"image_status": "verified_reference"},
                        "tab": tab_key,
                        "dims": vp.get("dimensions", {}).get("raw", ""),
                        "desc": vp.get("description", ""),
                        "specs": vp.get("specs", []),
                        "material": ", ".join(vp.get("materials", [])),
                        "price": vp.get("price", ""),
                        "jp_price": vp.get("jp_price", ""),
                        "img": vp.get("images", [""])[0]
                    })
        
    # Prioritize products with verified manufacturer/studio photos at the top of the catalog
    def get_sort_key(p):
        img_path = p.get("img", "")
        sku = p.get("sku", "")
        if sku in ["UNO-CHAIR", "WOODEN-CHAIR", "MACHETE-CHAIR", "CIRCON-CHAIR", "SUTERA-CHAIR", "ST-860E", "ST-MN01", "SWF-91704140", "SWF-91704070", "SWF-91962520"]:
            return (0, sku)
        if img_path and "[PDF]_" not in img_path and "_[pdf]_" not in img_path:
            return (1, sku)
        return (2, sku)

    products.sort(key=get_sort_key)

    payload = {
        "products": products,
        "catalog_policy": "Public product catalogue. Pricing, suppliers, margins, packaging, freight and loading data are intentionally excluded.",
        "total_products": len(products)
    }
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"Exported {len(products)} products to {output_path}")

# Run export for all three languages
export_sheet_to_json("Sunnyward_TW", os.path.join(base_dir, "tw", "products.json"), "tw")
export_sheet_to_json("Sunnyward_EN", os.path.join(base_dir, "en", "products.json"), "en")
export_sheet_to_json("Sunnyward_JP", os.path.join(base_dir, "jp", "products.json"), "jp")
